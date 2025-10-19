"""
üöÄ –ú–û–î–£–õ–¨ –î–õ–Ø –ü–†–û–§–ï–°–°–ò–û–ù–ê–õ–¨–ù–û–ô –ì–ï–ù–ï–†–ê–¶–ò–ò EXCEL –ó–ê–ö–ê–ó-–ù–ê–†–Ø–î–û–í
–§–ò–ù–ê–õ–¨–ù–ê–Ø –í–ï–†–°–ò–Ø –° –ò–°–ü–†–ê–í–õ–ï–ù–ù–´–ú –ë–ê–ì–û–ú –ö–û–ù–í–ï–†–¢–ê–¶–ò–ò –°–£–ú–ú–´
"""

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from num2words import num2words
import os
import logging
from typing import Dict, Any, Optional, Tuple, List

# ‚úÖ –ö–û–ù–ö–†–ï–¢–ù–´–ï –ò–°–ö–õ–Æ–ß–ï–ù–ò–Ø –î–õ–Ø EXCEL –ü–†–û–¶–ï–°–°–û–†–ê
class ExcelProcessingError(Exception):
    """–ë–∞–∑–æ–≤–∞—è –æ—à–∏–±–∫–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ Excel"""
    pass

class TemplateNotFoundError(ExcelProcessingError):
    """–û—à–∏–±–∫–∞: —à–∞–±–ª–æ–Ω –Ω–µ –Ω–∞–π–¥–µ–Ω"""
    pass

class ExcelGenerationError(ExcelProcessingError):
    """–û—à–∏–±–∫–∞ –≥–µ–Ω–µ—Ä–∞—Ü–∏–∏ Excel –¥–æ–∫—É–º–µ–Ω—Ç–∞"""
    pass

class FileSaveError(ExcelProcessingError):
    """–û—à–∏–±–∫–∞ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è —Ñ–∞–π–ª–∞"""
    pass

class AmountConversionError(ExcelProcessingError):
    """–û—à–∏–±–∫–∞ –∫–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏–∏ —Å—É–º–º—ã –≤ –ø—Ä–æ–ø–∏—Å—å"""
    pass

class FormattingError(ExcelProcessingError):
    """–û—à–∏–±–∫–∞ –ø—Ä–∏–º–µ–Ω–µ–Ω–∏—è —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è"""
    pass


class ExcelProcessor:
    def __init__(self):
        self.rate_per_hour = 2500

    def create_professional_order(self, session: Dict[str, Any], template_path: str, output_path: str) -> bool:
        """–°–û–ó–î–ê–ï–ú –ü–†–û–§–ï–°–°–ò–û–ù–ê–õ–¨–ù–´–ô –ó–ê–ö–ê–ó-–ù–ê–†–Ø–î –° –ß–ï–¢–ö–û–ô –°–¢–†–£–ö–¢–£–†–û–ô –ò –£–õ–£–ß–®–ï–ù–ù–û–ô –û–ë–†–ê–ë–û–¢–ö–û–ô –û–®–ò–ë–û–ö"""
        try:
            # –°–æ–∑–¥–∞–µ–º –Ω–æ–≤—É—é —Ä–∞–±–æ—á—É—é –∫–Ω–∏–≥—É
            wb = Workbook()
            ws = wb.active
            ws.title = "–ó–∞–∫–∞–∑-–Ω–∞—Ä—è–¥"
            
            # –ë–õ–û–ö 1: –®–ê–ü–ö–ê –î–û–ö–£–ú–ï–ù–¢–ê
            header_end_row = self._create_header_block(ws, session)
            
            # –ë–õ–û–ö 2: –†–ê–ë–û–¢–´
            works_start_row = header_end_row + 1  # –ù–∞—á–∏–Ω–∞–µ–º –ø–æ—Å–ª–µ —à–∞–ø–∫–∏
            works_end_row = self._create_works_block(ws, session, works_start_row)
            
            # –ë–õ–û–ö 3: –ú–ê–¢–ï–†–ò–ê–õ–´
            materials_start_row = works_end_row + 2  # –û—Ç—Å—Ç—É–ø –ø–æ—Å–ª–µ —Ä–∞–±–æ—Ç
            materials_end_row = self._create_materials_block(ws, session, materials_start_row)
            
            # –ë–õ–û–ö 4: –ò–¢–û–ì–ò –ò –°–£–ú–ú–ò–†–û–í–ê–ù–ò–ï
            totals_start_row = materials_end_row + 2  # –û—Ç—Å—Ç—É–ø –ø–æ—Å–ª–µ –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤
            totals_end_row = self._create_totals_block(ws, works_start_row, works_end_row, 
                                                     materials_start_row, materials_end_row, 
                                                     totals_start_row, session)
            
            # –ë–õ–û–ö 5: –ü–û–î–ü–ò–°–ò –ò –ö–û–ú–ú–ï–ù–¢–ê–†–ò–ò
            footer_start_row = totals_end_row + 2
            self._create_footer_block(ws, footer_start_row)
            
            # –ü–†–ò–ú–ï–ù–Ø–ï–ú –§–û–†–ú–ê–¢–ò–†–û–í–ê–ù–ò–ï
            self._apply_professional_formatting(ws, works_start_row, works_end_row, 
                                              materials_start_row, materials_end_row,
                                              totals_start_row, footer_start_row)
            
            # –°–æ—Ö—Ä–∞–Ω—è–µ–º —Ñ–∞–π–ª
            self._save_workbook_safely(wb, output_path)
            print(f"‚úÖ –ü—Ä–æ—Ñ–µ—Å—Å–∏–æ–Ω–∞–ª—å–Ω—ã–π Excel —Ñ–∞–π–ª —Å–æ–∑–¥–∞–Ω: {output_path}")
            return True
                
        except ExcelProcessingError:
            # –ü–µ—Ä–µ–±—Ä–∞—Å—ã–≤–∞–µ–º –Ω–∞—à–∏ –∫–æ–Ω–∫—Ä–µ—Ç–Ω—ã–µ –∏—Å–∫–ª—é—á–µ–Ω–∏—è
            raise
        except Exception as e:
            # –û–±–µ—Ä—Ç—ã–≤–∞–µ–º –Ω–µ–æ–∂–∏–¥–∞–Ω–Ω—ã–µ –æ—à–∏–±–∫–∏ –≤ –∫–æ–Ω–∫—Ä–µ—Ç–Ω–æ–µ –∏—Å–∫–ª—é—á–µ–Ω–∏–µ
            raise ExcelGenerationError(f"–ù–µ–æ–∂–∏–¥–∞–Ω–Ω–∞—è –æ—à–∏–±–∫–∞ –ø—Ä–∏ —Å–æ–∑–¥–∞–Ω–∏–∏ Excel: {e}") from e
    
    def _save_workbook_safely(self, wb: Workbook, output_path: str) -> None:
        """–ë–µ–∑–æ–ø–∞—Å–Ω–æ–µ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ —Ä–∞–±–æ—á–µ–π –∫–Ω–∏–≥–∏ —Å –æ–±—Ä–∞–±–æ—Ç–∫–æ–π –æ—à–∏–±–æ–∫"""
        try:
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º –¥–æ—Å—Ç—É–ø–Ω–æ—Å—Ç—å –¥–∏—Ä–µ–∫—Ç–æ—Ä–∏–∏
            output_dir = os.path.dirname(output_path)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            # –ü—ã—Ç–∞–µ–º—Å—è —Å–æ—Ö—Ä–∞–Ω–∏—Ç—å —Ñ–∞–π–ª
            wb.save(output_path)
            
            # –ü—Ä–æ–≤–µ—Ä—è–µ–º —á—Ç–æ —Ñ–∞–π–ª –¥–µ–π—Å—Ç–≤–∏—Ç–µ–ª—å–Ω–æ —Å–æ–∑–¥–∞–Ω
            if not os.path.exists(output_path):
                raise FileSaveError(f"–§–∞–π–ª –Ω–µ –±—ã–ª —Å–æ–∑–¥–∞–Ω: {output_path}")
                
            file_size = os.path.getsize(output_path)
            if file_size == 0:
                raise FileSaveError(f"–§–∞–π–ª —Å–æ–∑–¥–∞–Ω –ø—É—Å—Ç—ã–º: {output_path}")
                
        except PermissionError as e:
            raise FileSaveError(f"–ù–µ—Ç –ø—Ä–∞–≤ –¥–ª—è –∑–∞–ø–∏—Å–∏ —Ñ–∞–π–ª–∞: {output_path}") from e
        except OSError as e:
            raise FileSaveError(f"–û—à–∏–±–∫–∞ —Ñ–∞–π–ª–æ–≤–æ–π —Å–∏—Å—Ç–µ–º—ã –ø—Ä–∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–∏: {output_path}") from e
    
    def _create_header_block(self, ws, session: Dict[str, Any]) -> int:
        """–ë–õ–û–ö 1: –®–ê–ü–ö–ê –î–û–ö–£–ú–ï–ù–¢–ê –° –û–ë–†–ê–ë–û–¢–ö–û–ô –û–®–ò–ë–û–ö"""
        try:
            current_row = 1
            
            # –î–∞–Ω–Ω—ã–µ –∫–æ–º–ø–∞–Ω–∏–∏
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = "–ò–ù–î–ò–í–ò–î–£–ê–õ–¨–ù–´–ô –ü–†–ï–î–ü–†–ò–ù–ò–ú–ê–¢–ï–õ–¨ –ê–ô–†–ê–ü–ï–¢–Ø–ù –ö–†–ò–°–¢–ò–ù–ê –¢–ò–ì–†–ê–ù–û–í–ù–ê"
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = "–ò–ù–ù: 234206956031 –û–ì–†–ù–ò–ü: 321332800018501"
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = "600033, –í–ª–∞–¥–∏–º–∏—Ä—Å–∫–∞—è –æ–±–ª., –≥. –í–ª–∞–¥–∏–º–∏—Ä, —É–ª. –°—É—â–µ–≤—Å–∫–∞—è, –¥. 7, –∫–≤. 152"
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = "airanetan93@gmail.com +79190130122"
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # –ü—É—Å—Ç–∞—è —Å—Ç—Ä–æ–∫–∞
            current_row += 1
            
            # –ù–æ–º–µ—Ä –∑–∞–∫–∞–∑-–Ω–∞—Ä—è–¥–∞
            order_number = session.get('order_number', '000')
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = f"–ó–ê–ö–ê–ó ‚Äì –ù–ê–†–Ø–î ‚Ññ{order_number}"
            ws[f'B{current_row}'].font = Font(bold=True, size=14)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # –î–∞—Ç—ã
            date_str = session['date'].strftime('%d.%m.%Y')
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = f"–î–∞—Ç–∞ –∏ –≤—Ä–µ–º—è –ø—Ä–∏–µ–º–∞ –∑–∞–∫–∞–∑–∞: {date_str} –≥."
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = f"–î–∞—Ç–∞ –∏ –≤—Ä–µ–º—è –æ–∫–æ–Ω—á–∞–Ω–∏—è —Ä–∞–±–æ—Ç: {date_str} –≥."
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # –ó–∞–∫–∞–∑—á–∏–∫
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = "–ó–∞–∫–∞–∑—á–∏–∫"
            ws[f'B{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = "–ó–ê–û ¬´–ë—Ä–∏–¥–∂—Ç–∞—É–Ω –§—É–¥—Å¬ª"
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = "–ê–¥—Ä–µ—Å: 600026, –≥. –í–ª–∞–¥–∏–º–∏—Ä, —É–ª. –ö—É–π–±—ã—à–µ–≤–∞ –¥. 3"
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # –î–∞–Ω–Ω—ã–µ –∞–≤—Ç–æ–º–æ–±–∏–ª—è
            ws.merge_cells(f'B{current_row}:D{current_row}')
            ws[f'B{current_row}'] = f"–ú–∞—Ä–∫–∞, –º–æ–¥–µ–ª—å: Mercedes-Benz MP4"
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'E{current_row}'] = "–î–≤–∏–≥–∞—Ç–µ–ª—å ‚Ññ"
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'B{current_row}:D{current_row}')
            license_plate = session['license_plate']
            ws[f'B{current_row}'] = f"–ì–æ—Å—É–¥–∞—Ä—Å—Ç–≤–µ–Ω–Ω—ã–π —Ä–µ–≥. –Ω–æ–º–µ—Ä: {license_plate}"
            ws[f'B{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'E{current_row}'] = "–®–∞—Å—Å–∏ ‚Ññ"
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'B{current_row}:D{current_row}')
            ws[f'B{current_row}'] = "VIN"
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'E{current_row}'] = "–ö—É–∑–æ–≤ ‚Ññ"
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # –ó–∞–≥–æ–ª–æ–≤–æ–∫ —Ä–∞–∑–¥–µ–ª–∞ —Ä–∞–±–æ—Ç
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = f"–í—ã–ø–æ–ª–Ω–µ–Ω–Ω—ã–µ —Ä–∞–±–æ—Ç—ã –ø–æ –∑–∞–∫–∞–∑-–Ω–∞—Ä—è–¥—É ‚Ññ{order_number}"
            ws[f'B{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            print("‚úÖ –ë–ª–æ–∫ 1: –®–∞–ø–∫–∞ –¥–æ–∫—É–º–µ–Ω—Ç–∞ —Å–æ–∑–¥–∞–Ω–∞")
            return current_row  # –í–æ–∑–≤—Ä–∞—â–∞–µ–º –ø–æ—Å–ª–µ–¥–Ω—é—é —Å—Ç—Ä–æ–∫—É —à–∞–ø–∫–∏
            
        except Exception as e:
            raise ExcelGenerationError(f"–û—à–∏–±–∫–∞ —Å–æ–∑–¥–∞–Ω–∏—è —à–∞–ø–∫–∏ –¥–æ–∫—É–º–µ–Ω—Ç–∞: {e}") from e
    
    def _create_works_block(self, ws, session: Dict[str, Any], start_row: int) -> int:
        """–ë–õ–û–ö 2: –†–ê–ë–û–¢–´ –° –û–ë–†–ê–ë–û–¢–ö–û–ô –û–®–ò–ë–û–ö"""
        try:
            current_row = start_row
            
            # –ó–∞–≥–æ–ª–æ–≤–∫–∏ —Ç–∞–±–ª–∏—Ü–∏–∏ —Ä–∞–±–æ—Ç - –£–í–ï–õ–ò–ß–ò–í–ê–ï–ú –í–´–°–û–¢–£ –ò –ü–†–ò–ú–ï–ù–Ø–ï–ú –ü–ï–†–ï–ù–û–° –¢–ï–ö–°–¢–ê
            headers = ["‚Ññ", "–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ —Ä–∞–±–æ—Ç", "–ù–æ—Ä–º–∞ –≤—Ä–µ–º–µ–Ω–∏", "–ö–æ–ª-–≤–æ", "–°—Ç–æ–∏–º–æ—Å—Ç—å (—Ä—É–±.)", "–°—É–º–º–∞ (—Ä—É–±.)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                # –ü–ï–†–ï–ù–û–° –¢–ï–ö–°–¢–ê –î–õ–Ø –ó–ê–ì–û–õ–û–í–ö–û–í –¢–ê–ë–õ–ò–¶
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # –£–í–ï–õ–ò–ß–ò–í–ê–ï–ú –í–´–°–û–¢–£ –°–¢–†–û–ö–ò –ó–ê–ì–û–õ–û–í–ö–û–í –¢–ê–ë–õ–ò–¶–´ –†–ê–ë–û–¢
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            # –î–æ–±–∞–≤–ª—è–µ–º –≤—ã–±—Ä–∞–Ω–Ω—ã–µ —Ä–∞–±–æ—Ç—ã
            selected_works = session.get('selected_works', [])
            
            if not selected_works:
                # –ï—Å–ª–∏ —Ä–∞–±–æ—Ç –Ω–µ—Ç, –¥–æ–±–∞–≤–ª—è–µ–º —Å–æ–æ–±—â–µ–Ω–∏–µ
                ws.merge_cells(f"B{current_row}:F{current_row}")
                ws[f"B{current_row}"] = "–†–∞–±–æ—Ç—ã –Ω–µ –≤—ã–±—Ä–∞–Ω—ã"
                ws[f"B{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
            else:
                for i, (work_name, hours) in enumerate(selected_works, 1):
                    # –ü–æ—Ä—è–¥–∫–æ–≤—ã–π –Ω–æ–º–µ—Ä - –ü–û –¶–ï–ù–¢–†–£
                    ws.cell(row=current_row, column=1, value=i).alignment = Alignment(horizontal='center', vertical='center')
                    # –ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ —Ä–∞–±–æ—Ç—ã - –ü–û –õ–ï–í–û–ú–£
                    ws.cell(row=current_row, column=2, value=work_name).alignment = Alignment(horizontal='left', vertical='center')
                    # –ù–æ—Ä–º–∞ –≤—Ä–µ–º–µ–Ω–∏ - –ü–û –¶–ï–ù–¢–†–£
                    ws.cell(row=current_row, column=3, value=float(hours)).alignment = Alignment(horizontal='center', vertical='center')
                    # –ö–æ–ª-–≤–æ - –ü–û –¶–ï–ù–¢–†–£
                    ws.cell(row=current_row, column=4, value=1).alignment = Alignment(horizontal='center', vertical='center')
                    # –°—Ç–æ–∏–º–æ—Å—Ç—å - –ü–û –¶–ï–ù–¢–†–£
                    ws.cell(row=current_row, column=5, value=2500).alignment = Alignment(horizontal='center', vertical='center')
                    # –°—É–º–º–∞ - –ü–û –¶–ï–ù–¢–†–£
                    ws.cell(row=current_row, column=6, value=f"=C{current_row}*D{current_row}*E{current_row}").alignment = Alignment(horizontal='center', vertical='center')
                    
                    current_row += 1
            
            # –ò—Ç–æ–≥–æ —Ä–∞–±–æ—Ç - –ü–û –õ–ï–í–û–ú–£
            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws[f"B{current_row}"] = "–ò—Ç–æ–≥–æ —Ä–∞–±–æ—Ç—ã (—Ä—É–±.)"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='left', vertical='center')
            
            if selected_works:
                # –§–æ—Ä–º—É–ª–∞ —Å—É–º–º—ã —Ç–æ–ª—å–∫–æ –µ—Å–ª–∏ –µ—Å—Ç—å —Ä–∞–±–æ—Ç—ã
                first_data_row = start_row + 1
                last_data_row = current_row - 1
                ws[f"F{current_row}"] = f"=SUM(F{first_data_row}:F{last_data_row})"
            else:
                ws[f"F{current_row}"] = 0
                
            ws[f"F{current_row}"].font = Font(bold=True)
            ws[f"F{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"‚úÖ –ë–ª–æ–∫ 2: –†–∞–±–æ—Ç—ã —Å–æ–∑–¥–∞–Ω—ã ({len(selected_works)} –ø–æ–∑–∏—Ü–∏–π)")
            return current_row  # –í–æ–∑–≤—Ä–∞—â–∞–µ–º –ø–æ—Å–ª–µ–¥–Ω—é—é —Å—Ç—Ä–æ–∫—É –±–ª–æ–∫–∞ —Ä–∞–±–æ—Ç
            
        except Exception as e:
            raise ExcelGenerationError(f"–û—à–∏–±–∫–∞ —Å–æ–∑–¥–∞–Ω–∏—è –±–ª–æ–∫–∞ —Ä–∞–±–æ—Ç: {e}") from e
    
    def _create_materials_block(self, ws, session: Dict[str, Any], start_row: int) -> int:
        """–ë–õ–û–ö 3: –ú–ê–¢–ï–†–ò–ê–õ–´ - –¢–û–õ–¨–ö–û –í–´–ë–†–ê–ù–ù–´–ï –° –û–ë–†–ê–ë–û–¢–ö–û–ô –û–®–ò–ë–û–ö"""
        try:
            current_row = start_row
            
            order_number = session.get('order_number', '000')
            
            # –ó–∞–≥–æ–ª–æ–≤–æ–∫ —Ä–∞–∑–¥–µ–ª–∞ –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤
            ws.merge_cells(f"B{current_row}:F{current_row}")
            ws[f"B{current_row}"] = f"–†–∞—Å—Ö–æ–¥–Ω–∞—è –Ω–∞–∫–ª–∞–¥–Ω–∞—è –ø–æ –∑–∞–∫–∞–∑‚Äì–Ω–∞—Ä—è–¥—É ‚Ññ{order_number}"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # –ó–∞–≥–æ–ª–æ–≤–∫–∏ —Ç–∞–±–ª–∏—Ü—ã –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤ - –£–í–ï–õ–ò–ß–ò–í–ê–ï–ú –í–´–°–û–¢–£ –ò –ü–†–ò–ú–ï–ù–Ø–ï–ú –ü–ï–†–ï–ù–û–° –¢–ï–ö–°–¢–ê
            headers = ["‚Ññ", "–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ", "–ï–¥–∏–Ω–∏—Ü–∞ –∏–∑–º–µ—Ä–µ–Ω–∏—è", "–ö–æ–ª-–≤–æ", "–°—Ç–æ–∏–º–æ—Å—Ç—å (—Ä—É–±.)", "–°—É–º–º–∞ (—Ä—É–±.)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                # –ü–ï–†–ï–ù–û–° –¢–ï–ö–°–¢–ê –î–õ–Ø –ó–ê–ì–û–õ–û–í–ö–û–í –¢–ê–ë–õ–ò–¶
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # –£–í–ï–õ–ò–ß–ò–í–ê–ï–ú –í–´–°–û–¢–£ –°–¢–†–û–ö–ò –ó–ê–ì–û–õ–û–í–ö–û–í –¢–ê–ë–õ–ò–¶–´ –ú–ê–¢–ï–†–ò–ê–õ–û–í
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            # ‚úÖ –î–ò–ù–ê–ú–ò–ß–ï–°–ö–ò–ï –ú–ê–¢–ï–†–ò–ê–õ–´ - –¢–û–õ–¨–ö–û –í–´–ë–†–ê–ù–ù–´–ï
            selected_materials = session.get('selected_materials', [])
            materials_data = {
                "–í–î-40": ("—à—Ç.", 1, 375),
                "–ü–µ—Ä—á–∞—Ç–∫–∏": ("—à—Ç.", 1, 95), 
                "–°–º–∞–∑–∫–∞": ("—à—Ç.", 1, 210),
                "–î–∏—Å–∫ –æ—Ç—Ä–µ–∑–Ω–æ–π": ("—à—Ç.", 1, 120)
            }
            
            # –ï—Å–ª–∏ –Ω–∏—á–µ–≥–æ –Ω–µ –≤—ã–±—Ä–∞–Ω–æ - –∏—Å–ø–æ–ª—å–∑—É–µ–º –≤—Å–µ –º–∞—Ç–µ—Ä–∏–∞–ª—ã (–¥–ª—è –æ–±—Ä–∞—Ç–Ω–æ–π —Å–æ–≤–º–µ—Å—Ç–∏–º–æ—Å—Ç–∏)
            if not selected_materials:
                selected_materials = list(materials_data.keys())
            
            for i, material_name in enumerate(selected_materials, 1):
                if material_name in materials_data:
                    unit, qty, price = materials_data[material_name]
                    
                    # –ü–æ—Ä—è–¥–∫–æ–≤—ã–π –Ω–æ–º–µ—Ä - –ü–û –¶–ï–ù–¢–†–£
                    ws.cell(row=current_row, column=1, value=i).alignment = Alignment(horizontal='center', vertical='center')
                    # –ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ - –ü–û –õ–ï–í–û–ú–£
                    ws.cell(row=current_row, column=2, value=material_name).alignment = Alignment(horizontal='left', vertical='center')
                    # –ï–¥–∏–Ω–∏—Ü–∞ –∏–∑–º–µ—Ä–µ–Ω–∏—è - –ü–û –¶–ï–ù–¢–†–£
                    ws.cell(row=current_row, column=3, value=unit).alignment = Alignment(horizontal='center', vertical='center')
                    # –ö–æ–ª-–≤–æ - –ü–û –¶–ï–ù–¢–†–£
                    ws.cell(row=current_row, column=4, value=qty).alignment = Alignment(horizontal='center', vertical='center')
                    # –°—Ç–æ–∏–º–æ—Å—Ç—å - –ü–û –¶–ï–ù–¢–†–£
                    ws.cell(row=current_row, column=5, value=price).alignment = Alignment(horizontal='center', vertical='center')
                    # –°—É–º–º–∞ - –ü–û –¶–ï–ù–¢–†–£
                    ws.cell(row=current_row, column=6, value=f"=D{current_row}*E{current_row}").alignment = Alignment(horizontal='center', vertical='center')
                    
                    current_row += 1
            
            # –ò—Ç–æ–≥–æ –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤ - –ü–û –õ–ï–í–û–ú–£
            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws[f"B{current_row}"] = "–ò—Ç–æ–≥–æ –∑–∞–ø–∞—Å–Ω—ã–µ —á–∞—Å—Ç–∏ (—Ä—É–±.)"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='left', vertical='center')
            
            if selected_materials:
                first_data_row = start_row + 2
                last_data_row = current_row - 1
                ws[f"F{current_row}"] = f"=SUM(F{first_data_row}:F{last_data_row})"
            else:
                ws[f"F{current_row}"] = 0
                
            ws[f"F{current_row}"].font = Font(bold=True)
            ws[f"F{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"‚úÖ –ë–ª–æ–∫ 3: –ú–∞—Ç–µ—Ä–∏–∞–ª—ã —Å–æ–∑–¥–∞–Ω—ã ({len(selected_materials)} –ø–æ–∑–∏—Ü–∏–π)")
            return current_row
            
        except Exception as e:
            raise ExcelGenerationError(f"–û—à–∏–±–∫–∞ —Å–æ–∑–¥–∞–Ω–∏—è –±–ª–æ–∫–∞ –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤: {e}") from e
    
    def _create_totals_block(self, ws, works_start_row: int, works_end_row: int, materials_start_row: int, 
                           materials_end_row: int, start_row: int, session: Dict[str, Any]) -> int:
        """–ë–õ–û–ö 4: –ò–¢–û–ì–ò –ò –°–£–ú–ú–ò–†–û–í–ê–ù–ò–ï –° –û–ë–†–ê–ë–û–¢–ö–û–ô –û–®–ò–ë–û–ö"""
        try:
            current_row = start_row
            
            # –ó–∞–≥–æ–ª–æ–≤–æ–∫ –±–ª–æ–∫–∞ –∏—Ç–æ–≥–æ–≤ - –£–í–ï–õ–ò–ß–ò–í–ê–ï–ú –í–´–°–û–¢–£ –ò –ü–†–ò–ú–ï–ù–Ø–ï–ú –ü–ï–†–ï–ù–û–° –¢–ï–ö–°–¢–ê
            headers = ["‚Ññ", "–ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ", "", "", "", "–°—É–º–º–∞ (—Ä—É–±.)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                if header:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                    # –ü–ï–†–ï–ù–û–° –¢–ï–ö–°–¢–ê –î–õ–Ø –ó–ê–ì–û–õ–û–í–ö–û–í –¢–ê–ë–õ–ò–¶
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # –£–í–ï–õ–ò–ß–ò–í–ê–ï–ú –í–´–°–û–¢–£ –°–¢–†–û–ö–ò –ó–ê–ì–û–õ–û–í–ö–û–í –¢–ê–ë–õ–ò–¶–´ –ò–¢–û–ì–û–í
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            # –†–∞–±–æ—Ç–∞ (—Å—Å—ã–ª–∞–µ–º—Å—è –Ω–∞ —Å—Ç—Ä–æ–∫—É —Å –∏—Ç–æ–≥–æ–º —Ä–∞–±–æ—Ç)
            ws.cell(row=current_row, column=1, value=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=2, value="–†–∞–±–æ—Ç–∞").alignment = Alignment(horizontal='left', vertical='center')
            work_total_cell = f"F{works_end_row}"  # –°—Ç—Ä–æ–∫–∞ —Å –∏—Ç–æ–≥–æ–º —Ä–∞–±–æ—Ç
            ws.cell(row=current_row, column=6, value=f"={work_total_cell}").alignment = Alignment(horizontal='center', vertical='center')
            work_row = current_row
            current_row += 1
            
            # –ó–∞–ø–∞—Å–Ω—ã–µ —á–∞—Å—Ç–∏ (—Å—Å—ã–ª–∞–µ–º—Å—è –Ω–∞ —Å—Ç—Ä–æ–∫—É —Å –∏—Ç–æ–≥–æ–º –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤)
            ws.cell(row=current_row, column=1, value=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=2, value="–ó–∞–ø–∞—Å–Ω—ã–µ —á–∞—Å—Ç–∏").alignment = Alignment(horizontal='left', vertical='center')
            materials_total_cell = f"F{materials_end_row}"  # –°—Ç—Ä–æ–∫–∞ —Å –∏—Ç–æ–≥–æ–º –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤
            ws.cell(row=current_row, column=6, value=f"={materials_total_cell}").alignment = Alignment(horizontal='center', vertical='center')
            materials_row = current_row
            current_row += 1
            
            # –í—Å–µ–≥–æ –∫ –æ–ø–ª–∞—Ç–µ - –ü–û –õ–ï–í–û–ú–£
            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws[f"B{current_row}"] = "–í—Å–µ–≥–æ –∫ –æ–ø–ª–∞—Ç–µ (—Ä—É–±.)"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='left', vertical='center')
            ws[f"F{current_row}"] = f"=F{work_row}+F{materials_row}"
            ws[f"F{current_row}"].font = Font(bold=True)
            ws[f"F{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # –í—Å–µ–≥–æ –ø–æ –∑–∞–∫–∞–∑-–Ω–∞—Ä—è–¥—É - –ü–û –õ–ï–í–û–ú–£
            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws[f"B{current_row}"] = "–í—Å–µ–≥–æ –ø–æ –∑–∞–∫–∞–∑-–Ω–∞—Ä—è–¥—É:"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1
            
            # –°—É–º–º–∞ –ø—Ä–æ–ø–∏—Å—å—é - –†–ê–°–°–ß–ò–¢–´–í–ê–ï–ú –ü–†–ê–í–ò–õ–¨–ù–û
            # –ü–æ–ª—É—á–∞–µ–º —Å—É–º–º—É —Ä–∞–±–æ—Ç –∏ –º–∞—Ç–µ—Ä–∏–∞–ª–æ–≤ –∏–∑ session
            works_total = sum(hours for _, hours in session.get('selected_works', [])) * self.rate_per_hour
                
            selected_materials = session.get('selected_materials', [])
            materials_data = {
                "–í–î-40": 375,
                "–ü–µ—Ä—á–∞—Ç–∫–∏": 95, 
                "–°–º–∞–∑–∫–∞": 210,
                "–î–∏—Å–∫ –æ—Ç—Ä–µ–∑–Ω–æ–π": 120
            }
            
            if not selected_materials:
                selected_materials = list(materials_data.keys())
            
            materials_total = sum(materials_data.get(material, 0) for material in selected_materials)
            
            total_amount = works_total + materials_total
            amount_words = self._get_amount_in_words(total_amount)
            
            ws.merge_cells(f"B{current_row}:F{current_row}")
            ws[f"B{current_row}"] = amount_words
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='left', vertical='center')
            
            print("‚úÖ –ë–ª–æ–∫ 4: –ò—Ç–æ–≥–∏ –∏ —Å—É–º–º–∏—Ä–æ–≤–∞–Ω–∏–µ —Å–æ–∑–¥–∞–Ω—ã")
            return current_row
            
        except Exception as e:
            raise ExcelGenerationError(f"–û—à–∏–±–∫–∞ —Å–æ–∑–¥–∞–Ω–∏—è –±–ª–æ–∫–∞ –∏—Ç–æ–≥–æ–≤: {e}") from e
    
    def _create_footer_block(self, ws, start_row: int) -> None:
        """–ë–õ–û–ö 5: –ü–û–î–ü–ò–°–ò –ò –ö–û–ú–ú–ï–ù–¢–ê–†–ò–ò –° –û–ë–†–ê–ë–û–¢–ö–û–ô –û–®–ò–ë–û–ö"""
        try:
            current_row = start_row
            
            # –ü–æ–¥–ø–∏—Å–∏
            ws.merge_cells(f"B{current_row}:F{current_row}")
            ws[f"B{current_row}"] = "–ó–∞–∫–∞–∑—á–∏–∫________________                –ú–ü                          –ò—Å–ø–æ–ª–Ω–∏—Ç–µ–ª—å_______________       –ú–ü"
            ws[f"B{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 2
            
            # –ö–æ–º–º–µ–Ω—Ç–∞—Ä–∏–π
            ws.merge_cells(f"B{current_row}:F{current_row}")
            ws[f"B{current_row}"] = "–†–∞–±–æ—Ç—ã –≤—ã–ø–æ–ª–Ω–µ–Ω—ã —Å –∏—Å–ø–æ–ª—å–∑–æ–≤–∞–Ω–∏–µ–º –∑–∞–ø–∞—Å–Ω—ã—Ö —á–∞—Å—Ç–µ–π –∑–∞–∫–∞–∑—á–∏–∫–∞"
            ws[f"B{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            
            print("‚úÖ –ë–ª–æ–∫ 5: –ü–æ–¥–ø–∏—Å–∏ –∏ –∫–æ–º–º–µ–Ω—Ç–∞—Ä–∏–∏ —Å–æ–∑–¥–∞–Ω—ã")
            
        except Exception as e:
            raise ExcelGenerationError(f"–û—à–∏–±–∫–∞ —Å–æ–∑–¥–∞–Ω–∏—è –±–ª–æ–∫–∞ –ø–æ–¥–ø–∏—Å–µ–π: {e}") from e
    
    def _apply_professional_formatting(self, ws, works_start_row: int, works_end_row: int, 
                                    materials_start_row: int, materials_end_row: int,
                                    totals_start_row: int, footer_start_row: int) -> None:
        """–ü–†–ò–ú–ï–ù–Ø–ï–ú –ü–†–û–§–ï–°–°–ò–û–ù–ê–õ–¨–ù–û–ï –§–û–†–ú–ê–¢–ò–†–û–í–ê–ù–ò–ï –° –û–ë–†–ê–ë–û–¢–ö–û–ô –û–®–ò–ë–û–ö"""
        try:
            print("üé® –ü—Ä–∏–º–µ–Ω—è–µ–º –ø—Ä–æ—Ñ–µ—Å—Å–∏–æ–Ω–∞–ª—å–Ω–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ...")
            
            # –°–¢–ò–õ–¨ –ì–†–ê–ù–ò–¶
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # –®–ò–†–ò–ù–´ –ö–û–õ–û–ù–û–ö
            column_widths = {
                'A': 6,    # ‚Ññ
                'B': 45,   # –ù–∞–∏–º–µ–Ω–æ–≤–∞–Ω–∏–µ
                'C': 12,   # –ù–æ—Ä–º–∞ –≤—Ä–µ–º–µ–Ω–∏
                'D': 8,    # –ö–æ–ª-–≤–æ
                'E': 12,   # –°—Ç–æ–∏–º–æ—Å—Ç—å
                'F': 12    # –°—É–º–º–∞
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # –í–´–†–ê–í–ù–ò–í–ê–ù–ò–ï –ò –§–û–†–ú–ê–¢–ò–†–û–í–ê–ù–ò–ï –í–°–ï–• –Ø–ß–ï–ï–ö
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # –§–û–†–ú–ê–¢ –ß–ò–°–ï–õ
                        if cell.column in [5, 6] and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
            
            # –ì–†–ê–ù–ò–¶–´ –î–õ–Ø –¢–ê–ë–õ–ò–¶–´ –†–ê–ë–û–¢ (–≤–∫–ª—é—á–∞—è –∑–∞–≥–æ–ª–æ–≤–∫–∏ –∏ –∏—Ç–æ–≥–∏)
            for row in range(works_start_row, works_end_row + 1):
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border
            
            # –ì–†–ê–ù–ò–¶–´ –î–õ–Ø –¢–ê–ë–õ–ò–¶–´ –ú–ê–¢–ï–†–ò–ê–õ–û–í
            for row in range(materials_start_row, materials_end_row + 1):
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border
            
            # –ì–†–ê–ù–ò–¶–´ –î–õ–Ø –ë–õ–û–ö–ê –ò–¢–û–ì–û–í
            for row in range(totals_start_row, totals_start_row + 3):
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border
            
            print("‚úÖ –ü—Ä–æ—Ñ–µ—Å—Å–∏–æ–Ω–∞–ª—å–Ω–æ–µ —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏–µ –ø—Ä–∏–º–µ–Ω–µ–Ω–æ")
            
        except Exception as e:
            raise FormattingError(f"–û—à–∏–±–∫–∞ –ø—Ä–∏–º–µ–Ω–µ–Ω–∏—è —Ñ–æ—Ä–º–∞—Ç–∏—Ä–æ–≤–∞–Ω–∏—è: {e}") from e
    
    def _get_amount_in_words(self, amount: float) -> str:
        """‚úÖ –ò–°–ü–†–ê–í–õ–ï–ù–ù–´–ô –ú–ï–¢–û–î: –ö–æ–Ω–≤–µ—Ä—Ç–∏—Ä—É–µ—Ç —Å—É–º–º—É –≤ –ø—Ä–æ–ø–∏—Å—å –±–µ–∑ –æ—à–∏–±–æ–∫ –æ–∫—Ä—É–≥–ª–µ–Ω–∏—è"""
        try:
            # ‚úÖ –ò–°–ü–†–ê–í–õ–ï–ù–ò–ï: –ò—Å–ø–æ–ª—å–∑—É–µ–º —Ç–æ—á–Ω–æ–µ –≤—ã—á–∏—Å–ª–µ–Ω–∏–µ —á–µ—Ä–µ–∑ —Ü–µ–ª—ã–µ —á–∏—Å–ª–∞
            total_cents = round(amount * 100)  # –ü–µ—Ä–µ–≤–æ–¥–∏–º –≤—Å—é —Å—É–º–º—É –≤ –∫–æ–ø–µ–π–∫–∏
            rubles = total_cents // 100        # –¶–µ–ª–∞—è —á–∞—Å—Ç—å - —Ä—É–±–ª–∏
            kopecks = total_cents % 100        # –û—Å—Ç–∞—Ç–æ–∫ - –∫–æ–ø–µ–π–∫–∏
            
            if rubles < 0:
                raise AmountConversionError("–°—É–º–º–∞ –Ω–µ –º–æ–∂–µ—Ç –±—ã—Ç—å –æ—Ç—Ä–∏—Ü–∞—Ç–µ–ª—å–Ω–æ–π")
            
            rubles_words = num2words(rubles, lang='ru')
            
            # –û–ø—Ä–µ–¥–µ–ª—è–µ–º –ø—Ä–∞–≤–∏–ª—å–Ω–æ–µ —Å–∫–ª–æ–Ω–µ–Ω–∏–µ –¥–ª—è —Ä—É–±–ª–µ–π
            last_digit = rubles % 10
            last_two_digits = rubles % 100
            
            if 11 <= last_two_digits <= 19:
                ruble_word = "—Ä—É–±–ª–µ–π"
            elif last_digit == 1:
                ruble_word = "—Ä—É–±–ª—å"
            elif 2 <= last_digit <= 4:
                ruble_word = "—Ä—É–±–ª—è"
            else:
                ruble_word = "—Ä—É–±–ª–µ–π"
            
            # –§–æ—Ä–º–∞—Ç–∏—Ä—É–µ–º –∫–æ–ø–µ–π–∫–∏ –≤—Å–µ–≥–¥–∞ –¥–≤—É–º—è —Ü–∏—Ñ—Ä–∞–º–∏
            kopecks_text = f"{kopecks:02d}"
            
            result = f"{rubles_words.capitalize()} {ruble_word} {kopecks_text} –∫–æ–ø."
            print(f"üí∞ –°—É–º–º–∞ –ø—Ä–æ–ø–∏—Å—å—é: {result}")
            return result
            
        except Exception as e:
            error_msg = f"–û—à–∏–±–∫–∞ –∫–æ–Ω–≤–µ—Ä—Ç–∞—Ü–∏–∏ —Å—É–º–º—ã {amount}: {e}"
            print(f"‚ùå {error_msg}")
            raise AmountConversionError(error_msg) from e