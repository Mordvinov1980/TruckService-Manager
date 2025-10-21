"""
🚀 МОДУЛЬ ДЛЯ ПРОФЕССИОНАЛЬНОЙ ГЕНЕРАЦИИ EXCEL ЗАКАЗ-НАРЯДОВ
ФИНАЛЬНАЯ ВЕРСИЯ С ПОДДЕРЖКОЙ ШАБЛОНОВ ШАПОК
"""

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from num2words import num2words
import os
import logging
from typing import Dict, Any, Optional, Tuple, List
import json
import pathlib

# ✅ КОНКРЕТНЫЕ ИСКЛЮЧЕНИЯ ДЛЯ EXCEL ПРОЦЕССОРА
class ExcelProcessingError(Exception):
    """Базовая ошибка обработки Excel"""
    pass

class TemplateNotFoundError(ExcelProcessingError):
    """Ошибка: шаблон не найден"""
    pass

class ExcelGenerationError(ExcelProcessingError):
    """Ошибка генерации Excel документа"""
    pass

class FileSaveError(ExcelProcessingError):
    """Ошибка сохранения файла"""
    pass

class AmountConversionError(ExcelProcessingError):
    """Ошибка конвертации суммы в пропись"""
    pass

class FormattingError(ExcelProcessingError):
    """Ошибка применения форматирования"""
    pass

class HeaderTemplateManager:
    """Менеджер шаблонов шапок документов"""
    
    def __init__(self, templates_path: pathlib.Path):
        self.templates_path = templates_path
        self.templates = {}
        self._load_templates()
    
    def _load_templates(self) -> None:
        """Загрузка всех шаблонов шапок из папки"""
        try:
            # Создаем папку если не существует
            self.templates_path.mkdir(parents=True, exist_ok=True)
            
            template_files = list(self.templates_path.glob("*.json"))
            for template_file in template_files:
                try:
                    with open(template_file, 'r', encoding='utf-8') as f:
                        template_data = json.load(f)
                        self.templates[template_data['id']] = template_data
                    print(f"✅ Загружен шаблон: {template_data['name']}")
                except Exception as e:
                    print(f"❌ Ошибка загрузки шаблона {template_file}: {e}")
            
            print(f"✅ Всего загружено шаблонов шапок: {len(self.templates)}")
            
            # Если нет шаблонов, создаем базовые
            if not self.templates:
                self._create_default_templates()
                
        except Exception as e:
            print(f"❌ Ошибка загрузки шаблонов: {e}")

    def reload_templates(self) -> None:
        """ПЕРЕЗАГРУЗИТЬ ШАБЛОНЫ ИЗ ФАЙЛОВОЙ СИСТЕМЫ"""
        print("🔄 Перезагружаем шаблоны шапок...")
        self.templates = {}  # Очищаем кэш
        self._load_templates()  # Загружаем заново
        print(f"✅ Шаблоны перезагружены. Доступно: {len(self.templates)}")            
    
    def _create_default_templates(self) -> None:
        """Создание шаблонов по умолчанию"""
        default_templates = [
            {
                "id": "bridge_town",
                "name": "🏢 Бриджтаун Фудс",
                "customer": {
                    "company": "ЗАО «Бриджтаун Фудс»",
                    "address": "600026, г. Владимир, ул. Куйбышева д. 3"
                },
                "contractor": {
                    "company": "ИП Айрапетян Кристина Тиграновна",
                    "address": "600033, Владимирская обл., г. Владимир, ул. Сущевская, д. 7, кв. 152",
                    "inn": "234206956031",
                    "ogrnip": "321332800018501",
                    "email": "airanetan93@gmail.com",
                    "phone": "+79190130122"
                },
                "default_vehicle": "Mercedes-Benz MP4"
            },
            {
                "id": "company_a", 
                "name": "🏭 Компания А",
                "customer": {
                    "company": "ООО «Компания А»",
                    "address": "г. Москва, ул. Ленина д. 1"
                },
                "contractor": {
                    "company": "ИП Айрапетян Кристина Тиграновна",
                    "address": "600033, Владимирская обл., г. Владимир, ул. Сущевская, д. 7, кв. 152",
                    "inn": "234206956031", 
                    "ogrnip": "321332800018501",
                    "email": "airanetan93@gmail.com",
                    "phone": "+79190130122"
                },
                "default_vehicle": "Грузовой автомобиль"
            }
        ]
        
        for template_data in default_templates:
            template_file = self.templates_path / f"{template_data['id']}.json"
            try:
                with open(template_file, 'w', encoding='utf-8') as f:
                    json.dump(template_data, f, ensure_ascii=False, indent=2)
                self.templates[template_data['id']] = template_data
                print(f"✅ Создан шаблон по умолчанию: {template_data['name']}")
            except Exception as e:
                print(f"❌ Ошибка создания шаблона {template_data['id']}: {e}")
    
    def get_template(self, template_id: str) -> Dict[str, Any]:
        """Получить шаблон по ID"""
        return self.templates.get(template_id)
    
    def get_available_templates(self) -> List[Dict[str, str]]:
        """Получить список доступных шаблонов"""
        return [
            {'id': template_id, 'name': template_data['name']}
            for template_id, template_data in self.templates.items()
        ]

class ExcelProcessor:
    def __init__(self):
        self.rate_per_hour = 2500
        # ✅ ДОБАВЛЯЕМ МЕНЕДЖЕР ШАБЛОНОВ
        self.header_manager = HeaderTemplateManager(
            pathlib.Path("Шаблоны") / "header_templates"
        )
        # ✅ ДОБАВЛЯЕМ ОТЛАДКУ ЗАГРУЗКИ ШАБЛОНОВ
        print("🔍 DEBUG: Загружены шаблоны шапок:")
        templates = self.header_manager.get_available_templates()
        for template in templates:
            print(f"   - {template['name']} (ID: {template['id']})")
        print(f"🔍 DEBUG: Всего загружено шаблонов: {len(templates)}")

    def create_professional_order(self, session: Dict[str, Any], template_path: str, output_path: str) -> bool:
        """СОЗДАЕМ ПРОФЕССИОНАЛЬНЫЙ ЗАКАЗ-НАРЯД С ЧЕТКОЙ СТРУКТУРОЙ И УЛУЧШЕННОЙ ОБРАБОТКОЙ ОШИБОК"""
        try:
            # Создаем новую рабочую книгу
            wb = Workbook()
            ws = wb.active
            ws.title = "Заказ-наряд"
            
            # БЛОК 1: ШАПКА ДОКУМЕНТА
            header_end_row = self._create_header_block(ws, session)
            
            # БЛОК 2: РАБОТЫ
            works_start_row = header_end_row + 1  # Начинаем после шапки
            works_end_row = self._create_works_block(ws, session, works_start_row)
            
            # БЛОК 3: МАТЕРИАЛЫ
            materials_start_row = works_end_row + 2  # Отступ после работ
            materials_end_row = self._create_materials_block(ws, session, materials_start_row)
            
            # БЛОК 4: ИТОГИ И СУММИРОВАНИЕ
            totals_start_row = materials_end_row + 2  # Отступ после материалов
            totals_end_row = self._create_totals_block(ws, works_start_row, works_end_row, 
                                                     materials_start_row, materials_end_row, 
                                                     totals_start_row, session)
            
            # БЛОК 5: ПОДПИСИ И КОММЕНТАРИИ
            footer_start_row = totals_end_row + 2
            self._create_footer_block(ws, footer_start_row)
            
            # ПРИМЕНЯЕМ ФОРМАТИРОВАНИЕ
            self._apply_professional_formatting(ws, works_start_row, works_end_row, 
                                              materials_start_row, materials_end_row,
                                              totals_start_row, footer_start_row)
            
            # Сохраняем файл
            self._save_workbook_safely(wb, output_path)
            print(f"✅ Профессиональный Excel файл создан: {output_path}")
            return True
                
        except ExcelProcessingError:
            # Перебрасываем наши конкретные исключения
            raise
        except Exception as e:
            # Обертываем неожиданные ошибки в конкретное исключение
            raise ExcelGenerationError(f"Неожиданная ошибка при создании Excel: {e}") from e
    
    def _save_workbook_safely(self, wb: Workbook, output_path: str) -> None:
        """Безопасное сохранение рабочей книги с обработкой ошибок"""
        try:
            # Проверяем доступность директории
            output_dir = os.path.dirname(output_path)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir, exist_ok=True)
            
            # Пытаемся сохранить файл
            wb.save(output_path)
            
            # Проверяем что файл действительно создан
            if not os.path.exists(output_path):
                raise FileSaveError(f"Файл не был создан: {output_path}")
                
            file_size = os.path.getsize(output_path)
            if file_size == 0:
                raise FileSaveError(f"Файл создан пустым: {output_path}")
                
        except PermissionError as e:
            raise FileSaveError(f"Нет прав для записи файла: {output_path}") from e
        except OSError as e:
            raise FileSaveError(f"Ошибка файловой системы при сохранении: {output_path}") from e
    
    def _create_header_block(self, ws, session: Dict[str, Any]) -> int:
        """БЛОК 1: ШАПКА ДОКУМЕНТА С ПОДДЕРЖКОЙ ШАБЛОНОВ"""
        try:
            current_row = 1
            
            # ✅ ПОЛУЧАЕМ ВЫБРАННЫЙ ШАБЛОН ИЛИ ИСПОЛЬЗУЕМ ПО УМОЛЧАНИЮ
            template_id = session.get('header_template', 'bridge_town')
            template = self.header_manager.get_template(template_id)
            
            if not template:
                # Резервный шаблон если выбранный не найден
                template = self.header_manager.get_template('bridge_town')
            
            # ДАННЫЕ ИСПОЛНИТЕЛЯ (всегда одинаковые)
            contractor = template['contractor']
            customer = template['customer']
            
            # ШАПКА ДОКУМЕНТА С ДАННЫМИ ИЗ ШАБЛОНА
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = f"ИНДИВИДУАЛЬНЫЙ ПРЕДПРИНИМАТЕЛЬ {contractor['company'].split('ИП ')[1]}"
            ws[f'A{current_row}'].font = Font(bold=True, size=12)
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = f"ИНН: {contractor['inn']} ОГРНИП: {contractor['ogrnip']}"
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = contractor['address']
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:F{current_row}')
            ws[f'A{current_row}'] = f"{contractor['email']} {contractor['phone']}"
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Пустая строка
            current_row += 1
            
            # Номер заказ-наряда
            order_number = session.get('order_number', '000')
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = f"ЗАКАЗ – НАРЯД №{order_number}"
            ws[f'B{current_row}'].font = Font(bold=True, size=14)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Даты
            date_str = session['date'].strftime('%d.%m.%Y')
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = f"Дата и время приема заказа: {date_str} г."
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = f"Дата и время окончания работ: {date_str} г."
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # ЗАКАЗЧИК ИЗ ШАБЛОНА
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = "Заказчик"
            ws[f'B{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = customer['company']
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = f"Адрес: {customer['address']}"
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Данные автомобиля
            default_vehicle = template.get('default_vehicle', 'Автомобиль')
            ws.merge_cells(f'B{current_row}:D{current_row}')
            ws[f'B{current_row}'] = f"Марка, модель: {default_vehicle}"
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'E{current_row}'] = "Двигатель №"
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'B{current_row}:D{current_row}')
            license_plate = session['license_plate']
            ws[f'B{current_row}'] = f"Государственный рег. номер: {license_plate}"
            ws[f'B{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'E{current_row}'] = "Шасси №"
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            ws.merge_cells(f'B{current_row}:D{current_row}')
            ws[f'B{current_row}'] = "VIN"
            ws[f'B{current_row}'].alignment = Alignment(horizontal='left', vertical='center')
            ws[f'E{current_row}'] = "Кузов №"
            ws[f'E{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Заголовок раздела работ
            ws.merge_cells(f'B{current_row}:F{current_row}')
            ws[f'B{current_row}'] = f"Выполненные работы по заказ-наряду №{order_number}"
            ws[f'B{current_row}'].font = Font(bold=True)
            ws[f'B{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"✅ Блок 1: Шапка документа создана (шаблон: {template['name']})")
            return current_row
            
        except Exception as e:
            raise ExcelGenerationError(f"Ошибка создания шапки документа: {e}") from e
    
    def _create_works_block(self, ws, session: Dict[str, Any], start_row: int) -> int:
        """БЛОК 2: РАБОТЫ С ОБРАБОТКОЙ ОШИБОК"""
        try:
            current_row = start_row
            
            # Заголовки таблиции работ - УВЕЛИЧИВАЕМ ВЫСОТУ И ПРИМЕНЯЕМ ПЕРЕНОС ТЕКСТА
            headers = ["№", "Наименование работ", "Норма времени", "Кол-во", "Стоимость (руб.)", "Сумма (руб.)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                # ПЕРЕНОС ТЕКСТА ДЛЯ ЗАГОЛОВКОВ ТАБЛИЦ
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # УВЕЛИЧИВАЕМ ВЫСОТУ СТРОКИ ЗАГОЛОВКОВ ТАБЛИЦЫ РАБОТ
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            # Добавляем выбранные работы
            selected_works = session.get('selected_works', [])
            
            if not selected_works:
                # Если работ нет, добавляем сообщение
                ws.merge_cells(f"B{current_row}:F{current_row}")
                ws[f"B{current_row}"] = "Работы не выбраны"
                ws[f"B{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
            else:
                for i, (work_name, hours) in enumerate(selected_works, 1):
                    # Порядковый номер - ПО ЦЕНТРУ
                    ws.cell(row=current_row, column=1, value=i).alignment = Alignment(horizontal='center', vertical='center')
                    # Наименование работы - ПО ЛЕВОМУ
                    ws.cell(row=current_row, column=2, value=work_name).alignment = Alignment(horizontal='left', vertical='center')
                    # Норма времени - ПО ЦЕНТРУ
                    ws.cell(row=current_row, column=3, value=float(hours)).alignment = Alignment(horizontal='center', vertical='center')
                    # Кол-во - ПО ЦЕНТРУ
                    ws.cell(row=current_row, column=4, value=1).alignment = Alignment(horizontal='center', vertical='center')
                    # Стоимость - ПО ЦЕНТРУ
                    ws.cell(row=current_row, column=5, value=2500).alignment = Alignment(horizontal='center', vertical='center')
                    # Сумма - ПО ЦЕНТРУ
                    ws.cell(row=current_row, column=6, value=f"=C{current_row}*D{current_row}*E{current_row}").alignment = Alignment(horizontal='center', vertical='center')
                    
                    current_row += 1
            
            # Итого работ - ПО ЛЕВОМУ
            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws[f"B{current_row}"] = "Итого работы (руб.)"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='left', vertical='center')
            
            if selected_works:
                # Формула суммы только если есть работы
                first_data_row = start_row + 1
                last_data_row = current_row - 1
                ws[f"F{current_row}"] = f"=SUM(F{first_data_row}:F{last_data_row})"
            else:
                ws[f"F{current_row}"] = 0
                
            ws[f"F{current_row}"].font = Font(bold=True)
            ws[f"F{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"✅ Блок 2: Работы созданы ({len(selected_works)} позиций)")
            return current_row  # Возвращаем последнюю строку блока работ
            
        except Exception as e:
            raise ExcelGenerationError(f"Ошибка создания блока работ: {e}") from e
    
    def _create_materials_block(self, ws, session: Dict[str, Any], start_row: int) -> int:
        """БЛОК 3: МАТЕРИАЛЫ - ТОЛЬКО ВЫБРАННЫЕ С ОБРАБОТКОЙ ОШИБОК"""
        try:
            current_row = start_row
            
            order_number = session.get('order_number', '000')
            
            # Заголовок раздела материалов
            ws.merge_cells(f"B{current_row}:F{current_row}")
            ws[f"B{current_row}"] = f"Расходная накладная по заказ–наряду №{order_number}"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Заголовки таблицы материалов - УВЕЛИЧИВАЕМ ВЫСОТУ И ПРИМЕНЯЕМ ПЕРЕНОС ТЕКСТА
            headers = ["№", "Наименование", "Единица измерения", "Кол-во", "Стоимость (руб.)", "Сумма (руб.)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
                # ПЕРЕНОС ТЕКСТА ДЛЯ ЗАГОЛОВКОВ ТАБЛИЦ
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # УВЕЛИЧИВАЕМ ВЫСОТУ СТРОКИ ЗАГОЛОВКОВ ТАБЛИЦЫ МАТЕРИАЛОВ
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            # ✅ ДИНАМИЧЕСКИЕ МАТЕРИАЛЫ - ТОЛЬКО ВЫБРАННЫЕ
            selected_materials = session.get('selected_materials', [])
            materials_data = {
                "ВД-40": ("шт.", 1, 375),
                "Перчатки": ("шт.", 1, 95), 
                "Смазка": ("шт.", 1, 210),
                "Диск отрезной": ("шт.", 1, 120)
            }
            
            # Если ничего не выбрано - используем все материалы (для обратной совместимости)
            if not selected_materials:
                selected_materials = list(materials_data.keys())
            
            for i, material_name in enumerate(selected_materials, 1):
                if material_name in materials_data:
                    unit, qty, price = materials_data[material_name]
                    
                    # Порядковый номер - ПО ЦЕНТРУ
                    ws.cell(row=current_row, column=1, value=i).alignment = Alignment(horizontal='center', vertical='center')
                    # Наименование - ПО ЛЕВОМУ
                    ws.cell(row=current_row, column=2, value=material_name).alignment = Alignment(horizontal='left', vertical='center')
                    # Единица измерения - ПО ЦЕНТРУ
                    ws.cell(row=current_row, column=3, value=unit).alignment = Alignment(horizontal='center', vertical='center')
                    # Кол-во - ПО ЦЕНТРУ
                    ws.cell(row=current_row, column=4, value=qty).alignment = Alignment(horizontal='center', vertical='center')
                    # Стоимость - ПО ЦЕНТРУ
                    ws.cell(row=current_row, column=5, value=price).alignment = Alignment(horizontal='center', vertical='center')
                    # Сумма - ПО ЦЕНТРУ
                    ws.cell(row=current_row, column=6, value=f"=D{current_row}*E{current_row}").alignment = Alignment(horizontal='center', vertical='center')
                    
                    current_row += 1
            
            # Итого материалов - ПО ЛЕВОМУ
            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws[f"B{current_row}"] = "Итого запасные части (руб.)"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='left', vertical='center')
            
            if selected_materials:
                first_data_row = start_row + 2
                last_data_row = current_row - 1
                ws[f"F{current_row}"] = f"=SUM(F{first_data_row}:F{last_data_row})"
            else:
                ws[f"F{current_row}"] = 0
                
            ws[f"F{current_row}"].font = Font(bold=True)
            ws[f"F{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            
            print(f"✅ Блок 3: Материалы созданы ({len(selected_materials)} позиций)")
            return current_row
            
        except Exception as e:
            raise ExcelGenerationError(f"Ошибка создания блока материалов: {e}") from e
    
    def _create_totals_block(self, ws, works_start_row: int, works_end_row: int, materials_start_row: int, 
                           materials_end_row: int, start_row: int, session: Dict[str, Any]) -> int:
        """БЛОК 4: ИТОГИ И СУММИРОВАНИЕ С ОБРАБОТКОЙ ОШИБОК"""
        try:
            current_row = start_row
            
            # Заголовок блока итогов - УВЕЛИЧИВАЕМ ВЫСОТУ И ПРИМЕНЯЕМ ПЕРЕНОС ТЕКСТА
            headers = ["№", "Наименование", "", "", "", "Сумма (руб.)"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=current_row, column=col, value=header)
                if header:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="EEEEEE", end_color="EEEEEE", fill_type="solid")
                    # ПЕРЕНОС ТЕКСТА ДЛЯ ЗАГОЛОВКОВ ТАБЛИЦ
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # УВЕЛИЧИВАЕМ ВЫСОТУ СТРОКИ ЗАГОЛОВКОВ ТАБЛИЦЫ ИТОГОВ
            ws.row_dimensions[current_row].height = 30
            current_row += 1
            
            # Работа (ссылаемся на строку с итогом работ)
            ws.cell(row=current_row, column=1, value=1).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=2, value="Работа").alignment = Alignment(horizontal='left', vertical='center')
            work_total_cell = f"F{works_end_row}"  # Строка с итогом работ
            ws.cell(row=current_row, column=6, value=f"={work_total_cell}").alignment = Alignment(horizontal='center', vertical='center')
            work_row = current_row
            current_row += 1
            
            # Запасные части (ссылаемся на строку с итогом материалов)
            ws.cell(row=current_row, column=1, value=2).alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=current_row, column=2, value="Запасные части").alignment = Alignment(horizontal='left', vertical='center')
            materials_total_cell = f"F{materials_end_row}"  # Строка с итогом материалов
            ws.cell(row=current_row, column=6, value=f"={materials_total_cell}").alignment = Alignment(horizontal='center', vertical='center')
            materials_row = current_row
            current_row += 1
            
            # Всего к оплате - ПО ЛЕВОМУ
            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws[f"B{current_row}"] = "Всего к оплате (руб.)"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='left', vertical='center')
            ws[f"F{current_row}"] = f"=F{work_row}+F{materials_row}"
            ws[f"F{current_row}"].font = Font(bold=True)
            ws[f"F{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # Всего по заказ-наряду - ПО ЛЕВОМУ
            ws.merge_cells(f"B{current_row}:E{current_row}")
            ws[f"B{current_row}"] = "Всего по заказ-наряду:"
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='left', vertical='center')
            current_row += 1
            
            # Сумма прописью - РАССЧИТЫВАЕМ ПРАВИЛЬНО
            # Получаем сумму работ и материалов из session
            works_total = sum(hours for _, hours in session.get('selected_works', [])) * self.rate_per_hour
                
            selected_materials = session.get('selected_materials', [])
            materials_data = {
                "ВД-40": 375,
                "Перчатки": 95, 
                "Смазка": 210,
                "Диск отрезной": 120
            }
            
            if not selected_materials:
                selected_materials = list(materials_data.keys())
            
            materials_total = sum(materials_data.get(material, 0) for material in selected_materials)
            
            total_amount = works_total + materials_total
            amount_words = self._get_amount_in_words(total_amount)
            
            ws.merge_cells(f"B{current_row}:F{current_row}")
            ws[f"B{current_row}"] = amount_words
            ws[f"B{current_row}"].font = Font(bold=True)
            ws[f"B{current_row}"].alignment = Alignment(horizontal='left', vertical='center')
            
            print("✅ Блок 4: Итоги и суммирование созданы")
            return current_row
            
        except Exception as e:
            raise ExcelGenerationError(f"Ошибка создания блока итогов: {e}") from e
    
    def _create_footer_block(self, ws, start_row: int) -> None:
        """БЛОК 5: ПОДПИСИ И КОММЕНТАРИИ С ОБРАБОТКОЙ ОШИБОК"""
        try:
            current_row = start_row
            
            # Подписи
            ws.merge_cells(f"B{current_row}:F{current_row}")
            ws[f"B{current_row}"] = "Заказчик________________                МП                          Исполнитель_______________       МП"
            ws[f"B{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 2
            
            # Комментарий
            ws.merge_cells(f"B{current_row}:F{current_row}")
            ws[f"B{current_row}"] = "Работы выполнены с использованием запасных частей заказчика"
            ws[f"B{current_row}"].alignment = Alignment(horizontal='center', vertical='center')
            
            print("✅ Блок 5: Подписи и комментарии созданы")
            
        except Exception as e:
            raise ExcelGenerationError(f"Ошибка создания блока подписей: {e}") from e
    
    def _apply_professional_formatting(self, ws, works_start_row: int, works_end_row: int, 
                                    materials_start_row: int, materials_end_row: int,
                                    totals_start_row: int, footer_start_row: int) -> None:
        """ПРИМЕНЯЕМ ПРОФЕССИОНАЛЬНОЕ ФОРМАТИРОВАНИЕ С ОБРАБОТКОЙ ОШИБОК"""
        try:
            print("🎨 Применяем профессиональное форматирование...")
            
            # СТИЛЬ ГРАНИЦ
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # ШИРИНЫ КОЛОНОК
            column_widths = {
                'A': 6,    # №
                'B': 45,   # Наименование
                'C': 12,   # Норма времени
                'D': 8,    # Кол-во
                'E': 12,   # Стоимость
                'F': 12    # Сумма
            }
            
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            
            # ВЫРАВНИВАНИЕ И ФОРМАТИРОВАНИЕ ВСЕХ ЯЧЕЕК
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # ФОРМАТ ЧИСЕЛ
                        if cell.column in [5, 6] and isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'
            
            # ГРАНИЦЫ ДЛЯ ТАБЛИЦЫ РАБОТ (включая заголовки и итоги)
            for row in range(works_start_row, works_end_row + 1):
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border
            
            # ГРАНИЦЫ ДЛЯ ТАБЛИЦЫ МАТЕРИАЛОВ
            for row in range(materials_start_row, materials_end_row + 1):
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border
            
            # ГРАНИЦЫ ДЛЯ БЛОКА ИТОГОВ
            for row in range(totals_start_row, totals_start_row + 3):
                for col in range(1, 7):
                    ws.cell(row=row, column=col).border = thin_border
            
            print("✅ Профессиональное форматирование применено")
            
        except Exception as e:
            raise FormattingError(f"Ошибка применения форматирования: {e}") from e
    
    def _get_amount_in_words(self, amount: float) -> str:
        """✅ ИСПРАВЛЕННЫЙ МЕТОД: Конвертирует сумму в пропись без ошибок округления"""
        try:
            # ✅ ИСПРАВЛЕНИЕ: Используем точное вычисление через целые числа
            total_cents = round(amount * 100)  # Переводим всю сумму в копейки
            rubles = total_cents // 100        # Целая часть - рубли
            kopecks = total_cents % 100        # Остаток - копейки
            
            if rubles < 0:
                raise AmountConversionError("Сумма не может быть отрицательной")
            
            rubles_words = num2words(rubles, lang='ru')
            
            # Определяем правильное склонение для рублей
            last_digit = rubles % 10
            last_two_digits = rubles % 100
            
            if 11 <= last_two_digits <= 19:
                ruble_word = "рублей"
            elif last_digit == 1:
                ruble_word = "рубль"
            elif 2 <= last_digit <= 4:
                ruble_word = "рубля"
            else:
                ruble_word = "рублей"
            
            # Форматируем копейки всегда двумя цифрами
            kopecks_text = f"{kopecks:02d}"
            
            result = f"{rubles_words.capitalize()} {ruble_word} {kopecks_text} коп."
            print(f"💰 Сумма прописью: {result}")
            return result
            
        except Exception as e:
            error_msg = f"Ошибка конвертации суммы {amount}: {e}"
            print(f"❌ {error_msg}")
            raise AmountConversionError(error_msg) from e