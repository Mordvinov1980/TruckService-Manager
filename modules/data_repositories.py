"""
🚀 РЕПОЗИТОРИИ ДЛЯ РАБОТЫ С ДАННЫМИ
ПАТТЕРН REPOSITORY ДЛЯ АБСТРАКЦИИ ДОСТУПА К ДАННЫМ
"""

import pandas as pd
import pickle
import time
import pathlib
from typing import List, Tuple, Optional, Dict, Any
from abc import ABC, abstractmethod
import logging
import datetime

# ✅ БАЗОВЫЕ ИСКЛЮЧЕНИЯ ДЛЯ РЕПОЗИТОРИЕВ
class RepositoryError(Exception):
    """Базовая ошибка репозитория"""
    pass

class DataNotFoundError(RepositoryError):
    """Данные не найдены"""
    pass

class CacheError(RepositoryError):
    """Ошибка кэширования"""
    pass


# ✅ АБСТРАКТНЫЕ БАЗОВЫЕ КЛАССЫ
class BaseRepository(ABC):
    """Базовый репозиторий с кэшированием"""
    
    def __init__(self, cache_ttl: int = 3600):
        self.cache_ttl = cache_ttl
        self.logger = logging.getLogger('Repository')
    
    def _load_from_cache(self, cache_key: str, cache_file: pathlib.Path) -> Optional[Any]:
        """Загрузка данных из кэша"""
        try:
            if cache_file.exists():
                with open(cache_file, 'rb') as f:
                    cached_data = pickle.load(f)
                if time.time() - cached_data['timestamp'] < self.cache_ttl:
                    self.logger.info(f"✅ Загружено из кэша: {cache_key}")
                    return cached_data['data']
        except Exception as e:
            self.logger.warning(f"⚠️ Ошибка загрузки кэша {cache_key}: {e}")
        return None
    
    def _save_to_cache(self, data: Any, cache_key: str, cache_file: pathlib.Path) -> None:
        """Сохранение данных в кэш"""
        try:
            cache_file.parent.mkdir(parents=True, exist_ok=True)
            with open(cache_file, 'wb') as f:
                pickle.dump({'data': data, 'timestamp': time.time()}, f)
        except Exception as e:
            self.logger.warning(f"⚠️ Ошибка сохранения кэша {cache_key}: {e}")


class WorksRepository(BaseRepository):
    """Репозиторий для работы с работами"""
    
    @abstractmethod
    def get_works(self, section: str) -> List[Tuple[str, float]]:
        """Получить список работ для раздела"""
        pass
    
    @abstractmethod
    def get_works_count(self, section: str) -> int:
        """Получить количество работ в разделе"""
        pass


class MaterialsRepository(BaseRepository):
    """Репозиторий для работы с материалами"""
    
    @abstractmethod
    def get_materials(self) -> List[str]:
        """Получить список материалов"""
        pass
    
    @abstractmethod
    def get_material_price(self, material_name: str) -> float:
        """Получить цену материала"""
        pass


class AccountingRepository(BaseRepository):
    """Репозиторий для учета заказов"""
    
    @abstractmethod
    def save_order(self, session: Dict[str, Any], excel_filename: str, has_photos: str) -> bool:
        """Сохранить заказ в учет"""
        pass
    
    @abstractmethod
    def get_order_statistics(self, section: Optional[str] = None) -> Dict[str, Any]:
        """Получить статистику заказов"""
        pass


# ✅ КОНКРЕТНЫЕ РЕАЛИЗАЦИИ ДЛЯ EXCEL
class ExcelWorksRepository(WorksRepository):
    """Реализация репозитория работ для Excel"""
    
    def __init__(self, main_folder: pathlib.Path, sections_config: Dict[str, Any], cache_ttl: int = 3600):
        super().__init__(cache_ttl)
        self.main_folder = main_folder
        self.sections_config = sections_config
        
        # Стандартные работы на случай отсутствия файлов
        self.default_works = {
            'base': [
                ("Осмотр ТС", 0.4), ("Замена нижней накладки правой фары", 0.6),
                ("Замена верхней накладки правой фары", 0.8), ("Замена нижней накладки левой фары", 0.6),
                ("Замена верхней накладки левой фары", 0.8), ("Замена правой подножки", 1.4),
                ("Замена левой подножки", 1.4), ("Замена накладки правой подножки", 0.2),
                ("Замена накладки левой подножки", 0.2), ("Замена лючка правой подножки", 0.4),
            ]
        }
    
    def get_works(self, section: str) -> List[Tuple[str, float]]:
        """Получить работы для раздела с кэшированием"""
        if section not in self.sections_config:
            raise DataNotFoundError(f"Раздел не найден: {section}")
        
        # Проверяем кэш
        cache_key = f"{section}_works"
        cache_file = self.sections_config[section]['folder'] / "cache" / f"{cache_key}.pkl"
        
        cached_data = self._load_from_cache(cache_key, cache_file)
        if cached_data is not None:
            return cached_data
        
        # Загружаем из Excel
        works = self._load_works_from_excel(section)
        
        # Сохраняем в кэш
        self._save_to_cache(works, cache_key, cache_file)
        
        return works
    
    def get_works_count(self, section: str) -> int:
        """Получить количество работ в разделе"""
        works = self.get_works(section)
        return len(works)
    
    def _load_works_from_excel(self, section: str) -> List[Tuple[str, float]]:
        """Загрузка работ из Excel файла"""
        section_data = self.sections_config[section]
        works_file = section_data['works_file']
        
        try:
            # Ищем в корневой папке Шаблоны
            excel_path = self.main_folder / "Шаблоны" / works_file
            
            if not excel_path.exists():
                # Если нет в папке Шаблоны, пробуем прямо по имени файла
                excel_path = pathlib.Path(works_file)
            
            if excel_path.exists():
                df = pd.read_excel(excel_path)
                works = []
                valid_count = 0
                
                for idx, row in df.iterrows():
                    work_name = str(row.iloc[0]).strip()
                    work_hours = row.iloc[1]
                    
                    if work_name and work_name != 'nan' and pd.notna(work_hours):
                        try:
                            hours = float(work_hours)
                            works.append((work_name, hours))
                            valid_count += 1
                        except (ValueError, TypeError):
                            continue
                    else:
                        continue
                
                self.logger.info(f"✅ Загружено {valid_count} работ для раздела {section}")
                return works
            else:
                self.logger.warning(f"❌ Файл {works_file} не найден, используем стандартные работы")
                return self.default_works.get(section, [])
                
        except Exception as e:
            self.logger.error(f"❌ Ошибка загрузки Excel файла работ: {e}")
            return self.default_works.get(section, [])


class ExcelMaterialsRepository(MaterialsRepository):
    """Реализация репозитория материалов для Excel"""
    
    def __init__(self, main_folder: pathlib.Path, cache_ttl: int = 3600):
        super().__init__(cache_ttl)
        self.main_folder = main_folder
        
        # Цены материалов
        self.material_prices = {
            "ВД-40": 375,
            "Перчатки": 95,
            "Смазка": 210,
            "Диск отрезной": 120
        }
    
    def get_materials(self) -> List[str]:
        """Получить список материалов с кэшированием"""
        cache_key = "materials"
        cache_file = self.main_folder / "cache" / f"{cache_key}.pkl"
        
        cached_data = self._load_from_cache(cache_key, cache_file)
        if cached_data is not None:
            return cached_data
        
        # Загружаем из Excel
        materials = self._load_materials_from_excel()
        
        # Сохраняем в кэш
        self._save_to_cache(materials, cache_key, cache_file)
        
        return materials
    
    def get_material_price(self, material_name: str) -> float:
        """Получить цену материала"""
        return self.material_prices.get(material_name, 0)
    
    def _load_materials_from_excel(self) -> List[str]:
        """Загрузка материалов из Excel файла"""
        try:
            materials_path = self.main_folder / "Шаблоны" / "list_materials.xlsx"
            
            if not materials_path.exists():
                materials_path = pathlib.Path("list_materials.xlsx")
            
            if materials_path.exists():
                df = pd.read_excel(materials_path)
                materials = []
                valid_count = 0
                
                for idx, row in df.iterrows():
                    material_name = str(row.iloc[0]).strip()
                    
                    if material_name and material_name != 'nan':
                        materials.append(material_name)
                        valid_count += 1
                
                self.logger.info(f"✅ Загружено {valid_count} материалов")
                return materials
            else:
                self.logger.warning("❌ Файл list_materials.xlsx не найден, используем стандартные материалы")
                return list(self.material_prices.keys())
                
        except Exception as e:
            self.logger.error(f"❌ Ошибка загрузки материалов Excel: {e}")
            return list(self.material_prices.keys())


class ExcelAccountingRepository(AccountingRepository):
    """Реализация репозитория учета для Excel"""
    
    def __init__(self, main_folder: pathlib.Path, sections_config: Dict[str, Any], common_accounting_folder: pathlib.Path):
        super().__init__()
        self.main_folder = main_folder
        self.sections_config = sections_config
        self.common_accounting_folder = common_accounting_folder
        
        # Инициализируем файлы учета
        self._initialize_accounting_files()

    # ✅ ДОБАВИТЬ ЭТОТ МЕТОД ПРЯМО ЗДЕСЬ - после __init__ и перед save_order
    def _safe_dataframe_concat(self, df1: pd.DataFrame, df2: pd.DataFrame) -> pd.DataFrame:
        """Безопасное объединение DataFrame с обработкой пустых данных"""
        try:
            # Если оба DataFrame пустые
            if df1.empty and df2.empty:
                return pd.DataFrame()
            
            # Если первый DataFrame пустой
            if df1.empty:
                return df2
            
            # Если второй DataFrame пустой  
            if df2.empty:
                return df1
            
            # Оба не пустые - объединяем
            return pd.concat([df1, df2], ignore_index=True)
            
        except Exception as e:
            self.logger.error(f"Ошибка объединения DataFrame: {e}")
            # Возвращаем непустой DataFrame или создаем новый
            if not df1.empty:
                return df1
            elif not df2.empty:
                return df2
            else:
                return pd.DataFrame()

    def save_order(self, session: Dict[str, Any], excel_filename: str, has_photos: str) -> bool:
        """Сохранить заказ в учет"""
        try:
            section_id = session['section']
            
            # ✅ ВЫЧИСЛЯЕМ section_name ПЕРВЫМ ДЕЛОМ
            if section_id.startswith('custom_'):
                section_name = f"📁 {session['custom_list']}"
            else:
                section_name = self.sections_config[section_id]['name']
            
            # ✅ ПОТОМ перезаписываем section_id для учета
            if section_id.startswith('custom_'):
                section_id = 'base'
            
            # Сохраняем в раздельный учет
            section_accounting_file = self.sections_config[section_id]['folder'] / "Учет" / "учет_заказов.xlsx"
            
            # ✅ ИСПРАВЛЕННЫЕ КОЛОНКИ ДЛЯ РАЗДЕЛЬНОГО УЧЕТА (11 колонок)
            section_columns = [
                "ID", "Дата создания", "Время создания", "Номер ЗН", "Госномер", 
                "Исполнители", "Кол-во работ", "Общее время", "Файл Excel", "Файл черновика", "Фото добавлены"
            ]
            
            # Читаем существующие данные ИЛИ создаем новый DataFrame с правильными колонками
            try:
                df_section = pd.read_excel(section_accounting_file)
                # Проверяем, что все колонки присутствуют
                for col in section_columns:
                    if col not in df_section.columns:
                        df_section[col] = None
            except:
                df_section = pd.DataFrame(columns=section_columns)
            
            # Добавляем новую запись
            now = datetime.datetime.now()
            order_id = len(df_section) + 1 if not df_section.empty else 1
            selected_count = len(session['selected_works'])
            total_hours = sum(hours for _, hours in session['selected_works'])
            
            new_record = pd.DataFrame([{
                'ID': order_id,
                'Дата создания': session['date'].strftime('%d.%m.%Y'),
                'Время создания': now.strftime('%H:%M:%S'),
                'Номер ЗН': session.get('order_number', '000'),
                'Госномер': session['license_plate'],
                'Исполнители': session['workers'],
                'Кол-во работ': selected_count,
                'Общее время': total_hours,
                'Файл Excel': excel_filename,
                'Файл черновика': session.get('draft_filename', ''),
                'Фото добавлены': has_photos
            }])
            
            # Добавляем запись и сохраняем (исправленная конкатенация)
            df_section = pd.concat([df_section, new_record], ignore_index=True)
            df_section.to_excel(section_accounting_file, index=False)
            
            # ✅ ПРИМЕНЯЕМ АВТО-ФОРМАТИРОВАНИЕ ДЛЯ РАЗДЕЛЬНОГО УЧЕТА
            self._apply_accounting_formatting(section_accounting_file)
            
            # Сохраняем в общий учет
            common_accounting_file = self.common_accounting_folder / "главная_база.xlsx"
            
            # ✅ ИСПРАВЛЕННЫЕ КОЛОНКИ ДЛЯ ОБЩЕГО УЧЕТА (11 колонок)
            common_columns = [
                "ID", "Дата создания", "Время создания", "Раздел", "Номер ЗН", 
                "Госномер", "Исполнители", "Кол-во работ", "Общее время", "Файл Excel", "Фото добавлены"
            ]
            
            try:
                df_common = pd.read_excel(common_accounting_file)
                # Проверяем, что все колонки присутствуют
                for col in common_columns:
                    if col not in df_common.columns:
                        df_common[col] = None
            except:
                df_common = pd.DataFrame(columns=common_columns)
            
            new_common_record = pd.DataFrame([{
                'ID': len(df_common) + 1 if not df_common.empty else 1,
                'Дата создания': session['date'].strftime('%d.%m.%Y'),
                'Время создания': now.strftime('%H:%M:%S'),
                'Раздел': section_name,  # ✅ Теперь правильное имя раздела
                'Номер ЗН': session.get('order_number', '000'),
                'Госномер': session['license_plate'],
                'Исполнители': session['workers'],
                'Кол-во работ': selected_count,
                'Общее время': total_hours,
                'Файл Excel': excel_filename,
                'Фото добавлены': has_photos
            }])
            
            df_common = pd.concat([df_common, new_common_record], ignore_index=True)
            df_common.to_excel(common_accounting_file, index=False)
            
            # ✅ ПРИМЕНЯЕМ АВТО-ФОРМАТИРОВАНИЕ ДЛЯ ОБЩЕГО УЧЕТА
            self._apply_accounting_formatting(common_accounting_file)
            
            self.logger.info(f"✅ Заказ сохранен в учет: раздел {section_id}, ID {order_id}, фото: {has_photos}")
            return True
            
        except Exception as e:
            self.logger.error(f"❌ Ошибка сохранения в учет: {e}")
            return False
    
    def get_order_statistics(self, section: Optional[str] = None) -> Dict[str, Any]:
        """Получить статистику заказов (заглушка для будущей реализации)"""
        # TODO: Реализовать сбор статистики из файлов учета
        return {
            'total_orders': 0,
            'total_hours': 0,
            'total_revenue': 0
        }
    
    def _initialize_accounting_files(self) -> None:
        """Инициализация файлов учета"""
        try:
            for section_id in self.sections_config.keys():
                accounting_file = self.sections_config[section_id]['folder'] / "Учет" / "учет_заказов.xlsx"
                self._setup_section_accounting_file(accounting_file, section_id)
            
            common_accounting_file = self.common_accounting_folder / "главная_база.xlsx"
            self._setup_common_accounting_file(common_accounting_file)
            
        except Exception as e:
            raise RepositoryError(f"Ошибка инициализации учета: {e}") from e
    
    def _setup_section_accounting_file(self, accounting_file: pathlib.Path, section_id: str) -> None:
        """Создание файла учета для раздела"""
        if not accounting_file.exists():
            df = pd.DataFrame(columns=[
                "ID", "Дата создания", "Время создания", "Номер ЗН", "Госномер", 
                "Исполнители", "Кол-во работ", "Общее время", "Файл Excel", "Фото добавлены"
            ])
            df.to_excel(accounting_file, index=False)
    
    def _setup_common_accounting_file(self, accounting_file: pathlib.Path) -> None:
        """Создание общей базы данных"""
        if not accounting_file.exists():
            df = pd.DataFrame(columns=[
                "ID", "Дата создания", "Время создания", "Раздел", "Номер ЗН", 
                "Госномер", "Исполнители", "Кол-во работ", "Общее время", "Фото добавлены"
            ])
            df.to_excel(accounting_file, index=False)

    def _apply_accounting_formatting(self, file_path: pathlib.Path) -> None:
        """Применяет авто-форматирование к файлам учета"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Автоподбор ширины колонок
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Максимальная ширина 50
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # ✅ ВЫРАВНИВАНИЕ ПО ЦЕНТРУ ДЛЯ ВСЕХ ЯЧЕЕК
            for row in ws.iter_rows():
                for cell in row:
                    if cell.row == 1:  # Заголовки
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    else:
                        # ✅ ВСЕ ДАННЫЕ ПО ЦЕНТРУ
                        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Автоподбор высоты строки
                ws.row_dimensions[row[0].row].height = None
            
            # Закрепляем шапку
            ws.freeze_panes = 'A2'
            
            wb.save(file_path)
            print(f"✅ Применено авто-форматирование: {file_path.name}")
            
        except Exception as e:
            print(f"⚠️ Ошибка форматирования {file_path}: {e}")

# ✅ ФАБРИКА РЕПОЗИТОРИЕВ
class RepositoryFactory:
    """Фабрика для создания репозиториев"""
    
    @staticmethod
    def create_works_repository(main_folder: pathlib.Path, sections_config: Dict[str, Any]) -> WorksRepository:
        return ExcelWorksRepository(main_folder, sections_config)
    
    @staticmethod
    def create_materials_repository(main_folder: pathlib.Path) -> MaterialsRepository:
        return ExcelMaterialsRepository(main_folder)
    
    @staticmethod
    def create_accounting_repository(main_folder: pathlib.Path, sections_config: Dict[str, Any], common_accounting_folder: pathlib.Path) -> AccountingRepository:
        return ExcelAccountingRepository(main_folder, sections_config, common_accounting_folder)